from openpyxl import load_workbook
from interface_specs.G1int import G1int
from interface_specs.gig1_int import gig1_int
from interface_specs.tengig import ten_gig
from interface_specs.twentyfivegig import twenty_five_gig
from interface_specs.fortygig import forty_gig

workbook = load_workbook(filename="./config_files/excel_command_config.xlsx")

workbook.active = 0
sheet = workbook.active
g1ls = []
for g1int in sheet.iter_rows(min_row = 2, max_row = 49, values_only = True):
	interface = list(g1int)
	g1ls.append(G1int(interface))

workbook.active = 2
gig1_sheet = workbook.active
gig1_ls=[]
for gig1 in gig1_sheet.iter_rows(min_row = 2, max_row = 5, values_only = True):
	interfaces = list(gig1)
	gig1_ls.append(gig1_int(interfaces))
	print('done')

workbook.active = 3
tengig_sheet = workbook.active
gig10_ls=[]
for i in tengig_sheet.iter_rows(min_row = 2, max_row = 9, values_only = True):
	interface = list(i)
	gig10_ls.append(ten_gig(interface))

workbook.active = 4
gig25_sheet = workbook.active
gig25_ls=[]
for i in gig25_sheet.iter_rows(min_row = 2, max_row = 3, values_only = True):
	interface = list(i)
	gig25_ls.append(twenty_five_gig(interface))

workbook.active = 5
gig40_sheet = workbook.active
gig40_ls=[]
for i in gig40_sheet.iter_rows(min_row = 2, max_row = 3, values_only = True):
	interface = list(i)
	gig40_ls.append(forty_gig(interface))